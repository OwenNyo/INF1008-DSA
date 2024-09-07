import csv
import os
import requests
from flask import Flask, render_template, request, redirect, url_for, flash, session
import pandas as pd
from trafficFunctions import *
from folium import LayerControl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import time

app = Flask('__name__')
app.secret_key = 'd#5GKJJ8fj3kMLz.*dnsalnDAS&^%$#'
app.config['SESSION_TYPE'] = 'filesystem'

# Define TomTom API Details
TOMTOM_API_KEY = "Iv70lGtwRsTCFA2dCHgsnpJFKEe3XSzS"
TOMTOM_API_URL = "https://api.tomtom.com/traffic/services/5/incidentDetails"

# Folder Configs for Hotels, Bookings, ERPs, Maps
HOTELS_FOLDER = os.path.join(os.getcwd(), 'static', 'hotels')
BOOKINGS_FOLDER = os.path.join(os.getcwd(), 'static', 'bookings')
MAPS_FOLDER = os.path.join(os.getcwd(), 'static', 'route_maps')
ERP_FOLDER = os.path.join(os.getcwd(), 'static', 'ERP')
SHELL_FOLDER = os.path.join(os.getcwd(), 'static', 'Shell_Petrol')

app.config['HOTELS_FOLDER'] = HOTELS_FOLDER
app.config['BOOKINGS_FOLDER'] = BOOKINGS_FOLDER
app.config['MAPS_FOLDER'] = MAPS_FOLDER
app.config['ERP_FOLDER'] = ERP_FOLDER
app.config['SHELL_FOLDER'] = SHELL_FOLDER

# Global Variables
bus_size = 0
trip_time = ""
route_type = ""
unique_hotels_with_address = []
ERP_Passed = []
Traffic_Passed = []
earliest_time = 0
traffic_data = []


# Function to retrieve traffic data from TomTom API
def fetch_tomtom_traffic_data():
    bbox = "103.5709,1.2046,104.0635,1.4657"
    params = {
        "bbox": bbox,
        "fields": "{incidents{type,geometry{type,coordinates},properties{iconCategory}}}",
        "language": "en-GB",
        "categoryFilter": "1,7,8,9,14",
        "timeValidityFilter": "present",
        "key": TOMTOM_API_KEY
    }
    response = requests.get(TOMTOM_API_URL, params=params)
    response.raise_for_status()

    return response.json()


@app.route('/')
@app.route('/login')
def login():
    return render_template('login.html')


@app.route('/addHotel')
def add_hotel():
    return render_template('addHotel.html')


@app.route('/add_hotel', methods=['POST'])
def add_hotel_post():
    # Get hotel name and address from the form
    hotel_name = request.form['hotel-name']
    hotel_address = request.form['hotel-address']

    # Load the workbook
    wb_path = os.path.join(app.config['HOTELS_FOLDER'], 'Hotel_List.xlsx')
    wb = load_workbook(filename=wb_path)
    sheet = wb.active

    # Check for duplicates
    for row in sheet.iter_rows(values_only=True):
        if hotel_name in row:
            flash('Hotel already exists')
            return redirect(url_for('add_hotel'))

    # If no duplicates, add the hotel
    max_row = sheet.max_row
    sheet.cell(row=max_row + 1, column=1, value=hotel_name)
    sheet.cell(row=max_row + 1, column=2, value=hotel_address)

    # Convert address to longitude and latitude
    base_url = "https://maps.googleapis.com/maps/api/geocode/json"
    api_key = "AIzaSyAZ8zYtaQ-6vkkwFxeV1Ny1XQMSZiIXo6c"  # Replace with your actual API key
    params = {
        "address": hotel_address,
        "key": api_key
    }
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        data = response.json()
        if data['status'] == 'OK':
            latitude = data['results'][0]['geometry']['location']['lat']
            longitude = data['results'][0]['geometry']['location']['lng']
            sheet.cell(row=max_row + 1, column=3, value=latitude)
            sheet.cell(row=max_row + 1, column=4, value=longitude)

    # Save the workbook
    wb.save(os.path.join(app.config['HOTELS_FOLDER'], 'Hotel_List.xlsx'))

    # Flash a success message and redirect to home
    flash('Hotel added successfully')
    return redirect(url_for('add_hotel'))


@app.route('/guests')
def guests_hotels():
    wb = load_workbook(filename='static/hotels/Hotel_List.xlsx')
    sheet = wb.active
    hotels = [cell.value for cell in sheet['A'][1:] if
              cell.value is not None]  # assuming hotels are in the first column
    return render_template('guests.html', hotels=hotels)


@app.route('/guests')
def guests():
    return render_template('guests.html')


@app.route('/guests', methods=['POST'])
def submit_booking():
    username = request.form.get('username')
    contact_number = request.form.get('contact_number')
    select_date = request.form.get('select_date')
    select_time = request.form.get('select_time')
    name_hotel = request.form.get('name_hotel')

    # Format the date to %d/%m/%Y
    formatted_date = datetime.strptime(select_date, '%Y-%m-%d').strftime('%d/%m/%Y')

    # Define the path to the CSV file
    csv_path = os.path.join(app.config['BOOKINGS_FOLDER'], 'guests_bookings.csv')

    # Check if the CSV file exists
    file_exists = os.path.isfile(csv_path)

    # Open the CSV file in append mode
    with open(csv_path, mode='a', newline='') as file:
        writer = csv.writer(file)

        # Write the header if the file is new
        if not file_exists:
            writer.writerow(['Username', 'Contact Number', 'Date', 'Time', 'HotelName'])

        # Add the booking information
        writer.writerow([username, contact_number, formatted_date, select_time, name_hotel])

    # Flash a success message and redirect to home
    flash('Booking submitted successfully')
    return redirect(url_for('view_booking'))


@app.route('/trafficUpdates')
def traffic_updates():
    global traffic_data
    traffic_data = fetch_tomtom_traffic_data()
    return render_template('trafficUpdates.html', traffic_data=traffic_data)


@app.route('/guest_login', methods=['POST'])
def guest_login():
    # Process the login form data if needed
    username = request.form['username']
    password = request.form['password']
    session['username'] = username
    return redirect(url_for('guests'))


@app.route('/operator_login', methods=['POST'])
def operator_login():
    # Process the login form data if needed
    username = request.form['username']
    password = request.form['password']

    return redirect(url_for('traffic_updates'))


@app.route('/sortBookings')
def sort_bookings():

        # Data file path
        booking_file_path = os.path.join(app.config['BOOKINGS_FOLDER'], 'guests_bookings.csv')
        sorted_file_path = os.path.join(app.config['BOOKINGS_FOLDER'], 'Bookings_Sorted.csv')

        # Read the data from the Excel file
        bookings_df = pd.read_csv(booking_file_path)

        # Convert the 'Date' column to date format (remove time)
        bookings_df['Date'] = pd.to_datetime(bookings_df['Date'], format='%d/%m/%Y').dt.date

        # Convert the 'Time' column to time format
        bookings_df['Time'] = pd.to_datetime(bookings_df['Time']).dt.time

        # Get the date for tomorrow
        tomorrow = datetime.now().date() + timedelta(days=1)

        # Filter the data for tomorrow's date
        bookings_df = bookings_df[bookings_df['Date'] == tomorrow]

        # Check if the sorted file is empty
        if os.path.exists(sorted_file_path) and os.path.getsize(sorted_file_path) > 0:

            # Read the existing data from the sorted file
            existing_data = pd.read_csv(sorted_file_path)

            # Check if there is only the header (no data)
            if len(existing_data) <= 1:
                # Proceed with sorting as there is only the header or the file is empty
                bookings_df = bookings_df.sort_values(by=['Time', 'HotelName'])

                # Save the sorted data to the sorted CSV file
                bookings_df.to_csv(sorted_file_path, index=False)
                return redirect(url_for('job_assignment', success=True))
            else:
                # Skip sorting as there is data besides the header
                return redirect(url_for('job_assignment', success=False))

@app.route('/jobAssignment')
def job_assignment():
    return render_template('jobAssignment.html')


@app.route('/jobAssignment_assign')
def job_assignment_assign():
    global bus_size, unique_hotels_with_address, earliest_time, trip_time

    # Path to customer booking Excel file
    booking_file_path = os.path.join(app.config['BOOKINGS_FOLDER'], 'Bookings_Sorted.csv')

    # Path to the hotel list Excel file
    hotel_list_file_path = os.path.join(app.config['HOTELS_FOLDER'], 'Hotel_List.xlsx')

    # Read the data from the CSV file
    bookings_df = pd.read_csv(booking_file_path)

    # Check if there are any records in bookings_df
    if len(bookings_df) > 0:
        # Attempt to get the earliest time
        earliest_time = bookings_df['Time'].iloc[0]  # Access the first record's 'Time'

        # Used to pass and determine which ERP is active
        trip_time = earliest_time

        # Filter records for the earliest time
        earliest_records = bookings_df[bookings_df['Time'] == earliest_time]

        # Get the first 30 records or less if there are not enough
        records_to_take = earliest_records.head(30)

        # Determine the bus size based on the number of records pulled
        num_records = len(records_to_take)
        if num_records > 30:
            bus_size = 30
        elif 20 < num_records <= 30:
            bus_size = 30
        elif 10 < num_records <= 20:
            bus_size = 20
        else:
            bus_size = 10

        # Store unique hotel names
        unique_hotels = records_to_take[['HotelName']].drop_duplicates().to_dict('records')

        # Read the hotel list Excel file
        hotel_list_df = pd.read_excel(hotel_list_file_path)

        # Create a dictionary for hotel name to address mapping
        hotel_address_dict = pd.Series(hotel_list_df.HotelAddress.values,
                                       index=hotel_list_df.HotelName.str.strip()).to_dict()

        unique_hotels_with_address.clear()
        # Prepare the list of unique hotels with addresses
        for hotel in unique_hotels:
            hotel_name = hotel['HotelName']
            address = hotel_address_dict.get(hotel_name, 'Address Not Found')
            unique_hotels_with_address.append({'HotelName': hotel_name, 'HotelAddress': address})

        # Update the DataFrame by removing the processed records
        # remaining_records = bookings_df.drop(records_to_take.index)

        # Save the updated DataFrame back to the Excel file
        # remaining_records.to_csv(booking_file_path, index=False)

    else:
        # Handle case where there are no more records in bookings_df
        earliest_time = None  # or set a default value
        bus_size = 0  # or handle bus size as per your requirements
        unique_hotels_with_address.clear()  # clear the list of hotels

    # Render the template with the appropriate data
    return render_template('jobAssignment.html', unique_hotels=unique_hotels_with_address,
                           earliest_time=earliest_time, bus_size=bus_size)


@app.route('/jobAssignment_map', methods=['GET', 'POST'])
def job_assignment_map():
    # Global Variables
    global ERP_Passed, bus_size, trip_time, route_type, unique_hotels_with_address, earliest_time, traffic_data
    global Traffic_Passed

    # Static Variables
    total_erp_charge = 0
    route_distance = []
    Plotted_Map = None
    nearest_shell_data = None
    petrol_low = None
    start_time = time.time()
    traffic_coordinates_list = extract_traffic_data_coordinates(traffic_data)

    # Path to the Excel file
    hotel_list_file_path = os.path.join(HOTELS_FOLDER, 'Hotel_List.xlsx')
    erp_file_path = os.path.join(ERP_FOLDER, 'erp_coordinates.csv')
    shell_file_path = os.path.join(SHELL_FOLDER, 'Shell.csv')

    # Load the ERP data from the CSV file
    loaded_erp_df = pd.read_csv(erp_file_path)
    loaded_shell_df = pd.read_csv(shell_file_path)

    # Create a dictionary to store ERP coordinates
    ERP_COORDINATES = {}
    SHELL_COORDINATES = {}
    active_marker = folium.FeatureGroup(name='Active ERP Markers', show=False)
    inactive_marker = folium.FeatureGroup(name='Inactive ERP Markers', show=False)
    traffic_marker = folium.FeatureGroup(name='Traffic Data Markers', show=False)
    shell_marker = folium.FeatureGroup(name='Petrol Station Markers', show=False)

    # Populate ERP & Traffic markers
    for erp_name, row in loaded_erp_df.iterrows():
        erp_name = row['ERP Name']
        ERP_COORDINATES[erp_name] = {
            "coords": (row['Latitude'], row['Longitude']),
            "times": row['Times'],
            "charges": row['Charges']
        }
        active, active_time_range = is_erp_active(row['Times'],
                                                  trip_time)
        icon_color = 'green' if active else 'red'
        marker_layer = active_marker if active else inactive_marker

        # Filter charges and times if active
        if active:
            times_list = row['Times'].split(", ")
            charges_list = row['Charges'].split(", ")
            if active_time_range in times_list:
                active_index = times_list.index(active_time_range)
                if active_index < len(charges_list):
                    active_charge = charges_list[active_index]
                    popup_text = f"{erp_name}<br>Charge: {active_charge}<br>Time: {active_time_range}"
        else:
            popup_text = f"{erp_name}<br>Charges: {row['Charges']}<br>Times: {row['Times']}"

        folium.Marker(
            location=(row['Latitude'], row['Longitude']),
            popup=popup_text,
            icon=folium.Icon(color=icon_color, icon='info-sign')
        ).add_to(marker_layer)

    for lat, lon in traffic_coordinates_list:
        folium.Marker(
            location=[lat, lon],
            icon=folium.Icon(color='orange', icon='info-sign')
        ).add_to(traffic_marker)

    for idx, row in loaded_shell_df.iterrows():
        shell_name = row['station_name']  # Assuming there's a column for the name
        SHELL_COORDINATES[shell_name] = {
            "coords": (row['station_lat'], row['station_long'])
        }
        popup_text = f"{shell_name}<br>Address: {row['station_add']}"  # Adjust based on available data

        folium.Marker(
            location=(row['station_lat'], row['station_long']),
            popup=popup_text,
            icon=folium.Icon(color='gray', icon='info-sign')  # Color and icon can be customized
        ).add_to(shell_marker)

    if request.method == 'POST':
        session['petrol_low'] = 'petrolLowCheckbox' in request.form

    petrol_low = session.get('petrol_low', False)

    # Create projected graph and map
    start_location = "Changi Airport T3"
    coordinates = load_coordinates_from_csv(hotel_list_file_path, unique_hotels_with_address)
    Map = create_map(coordinates, start_location)
    start_coordinates = coordinates[start_location]['coords']
    Graph = download_road_network(start_coordinates)
    nodes = find_nearest_nodes(Graph, coordinates)

    # Check if the route type is selected
    if route_type:
        print(route_type)
    else:
        route_type = "shortest_distance"
        print(route_type)

    if route_type == "shortest_distance":
        # Determine the shortest path for first route
        try:
            route_shortest = shortest_path_route(Graph, nodes, start_location)

            if petrol_low:
                last_hotel_name = route_shortest[-1]
                last_hotel_coords = coordinates[last_hotel_name]['coords']
                nearest_shell, shell_coords, nearest_node, nearest_shell_data, shell_distance = find_nearest_shell(last_hotel_coords[0], last_hotel_coords[1], loaded_shell_df, Graph)
                route_shortest.append(nearest_shell)
                coordinates[nearest_shell] = {'coords': shell_coords}
                nodes[nearest_shell] = nearest_node
                print(f"Nearest Shell Station: {nearest_shell} at {shell_distance:.2f} km")

            print("Route Order:", route_shortest)
        except ValueError as e:
            print(e)
            route_shortest = []
        # Calculate and plot the first route
        # ERP & Traffic_Passed Array only initialized once here to get the stats for the first route
        pri_route, route_distance, ERP_Passed, Traffic_Passed, Shell_Passed = calculate_full_route(Graph, nodes
                                                                                                   , route_shortest
                                                                                                   , ERP_COORDINATES
                                                                                                   , traffic_coordinates_list
                                                                                                   , nearest_shell_data
                                                                                                   , trip_time)

        Plotted_Map = plot_route_on_map(Map, Graph, pri_route, ERP_COORDINATES, ERP_Passed, Traffic_Passed, Shell_Passed, color="green")

        print("Plotted Shortest")
    elif route_type == "fastest_route":
        Graph_no_traffic = avoid_traffic(Graph, traffic_coordinates_list)
        try:
            route_no_traffic = shortest_path_route(Graph_no_traffic, nodes, start_location)

            if petrol_low:
                last_hotel_name = route_no_traffic[-1]
                last_hotel_coords = coordinates[last_hotel_name]['coords']
                nearest_shell, shell_coords, nearest_node, nearest_shell_data, shell_distance = find_nearest_shell(last_hotel_coords[0], last_hotel_coords[1], loaded_shell_df, Graph)
                route_no_traffic.append(nearest_shell)
                coordinates[nearest_shell] = {'coords': shell_coords}
                nodes[nearest_shell] = nearest_node
                print(f"Nearest Shell Station: {nearest_shell} at {shell_distance:.2f} km")

            print("Route Order:", route_no_traffic)
        except ValueError as e:
            print(e)
            route_no_traffic = []

        # Calculate and plot the no erp route
        no_traffic_route, route_distance, ERP_Passed, Traffic_Passed, Shell_Passed = calculate_full_route(Graph, nodes
                                                                                                   , route_no_traffic
                                                                                                   , ERP_COORDINATES
                                                                                                   , traffic_coordinates_list
                                                                                                   , nearest_shell_data
                                                                                                   , trip_time)
        Plotted_Map = plot_route_on_map(Map, Graph_no_traffic, no_traffic_route,
                                        ERP_COORDINATES, ERP_Passed, Traffic_Passed, Shell_Passed, color="green")

        print("plotted avoid traffic route")
    elif route_type == "fastest_avoid_erp":
        # Avoid both ERP and traffic by combining the graph transformations
        Graph_no_erp = avoid_erp(Graph, ERP_COORDINATES)
        Graph_no_erp_and_traffic = avoid_traffic(Graph_no_erp, traffic_coordinates_list)

        try:
            route_order = shortest_path_route(Graph_no_erp_and_traffic, nodes, start_location)
            if petrol_low:
                last_hotel_name = route_order[-1]
                last_hotel_coords = coordinates[last_hotel_name]['coords']
                nearest_shell, shell_coords, nearest_node, nearest_shell_data, shell_distance = find_nearest_shell(last_hotel_coords[0], last_hotel_coords[1], loaded_shell_df, Graph)
                route_order.append(nearest_shell)
                coordinates[nearest_shell] = {'coords': shell_coords}
                nodes[nearest_shell] = nearest_node
                print(f"Nearest Shell Station: {nearest_shell} at {shell_distance:.2f} km")
            print("Route Order:", route_order)
        except ValueError as e:
            print(e)
            route_order = []

        # Calculate and plot the combined avoidance route
        combined_route, route_distance, ERP_Passed, Traffic_Passed, Shell_Passed = calculate_full_route(Graph_no_erp_and_traffic,
                                                                    nodes, route_order,
                                                                    ERP_COORDINATES,
                                                                    traffic_coordinates_list, nearest_shell_data,
                                                                    trip_time)
        print(ERP_Passed)
        Plotted_Map = plot_route_on_map(Map, Graph_no_erp_and_traffic, combined_route, ERP_COORDINATES, ERP_Passed,
                                        Traffic_Passed, Shell_Passed, color="green")
        print("plotted fastest + avoid erp route")
    else:
        print("plotting avoid erp")
        # Plot 2nd Route (Avoids ERP)
        Graph_no_erp = avoid_erp(Graph, ERP_COORDINATES)
        try:
            route_no_erp = shortest_path_route(Graph_no_erp, nodes, start_location)
            if petrol_low:
                last_hotel_name = route_no_erp[-1]
                last_hotel_coords = coordinates[last_hotel_name]['coords']
                nearest_shell, shell_coords, nearest_node, nearest_shell_data, shell_distance = find_nearest_shell(last_hotel_coords[0], last_hotel_coords[1], loaded_shell_df, Graph)
                route_no_erp.append(nearest_shell)
                coordinates[nearest_shell] = {'coords': shell_coords}
                nodes[nearest_shell] = nearest_node
                print(f"Nearest Shell Station: {nearest_shell} at {shell_distance:.2f} km")
            print("Route Order:", route_no_erp)
        except ValueError as e:
            print(e)
            route_no_erp = []

        # Calculate and plot the no erp route
        no_erp_route, route_distance, ERP_Passed, Traffic_Passed, Shell_Passed = calculate_full_route(Graph_no_erp, nodes, route_no_erp,
                                                                  ERP_COORDINATES, traffic_coordinates_list, nearest_shell_data, trip_time)
        Plotted_Map = plot_route_on_map(Map, Graph_no_erp, no_erp_route, ERP_COORDINATES, ERP_Passed, Traffic_Passed, Shell_Passed,
                                        color="green")
        print("plotted avoid erp")

    # Calculate Total Distances
    total_dist = sum(distance[2] for distance in route_distance)

    # Set the fixed cost per litre
    DieselCost_per_litre = 2.52

    # Determine fuel consumption based on bus type
    if bus_size == 10:
        FuelConsumption_per_100km = 15
    elif bus_size == 20:
        FuelConsumption_per_100km = 20
    elif bus_size == 30:
        FuelConsumption_per_100km = 25
    else:
        FuelConsumption_per_100km = 0  # Default case if no valid selection is made

    # Calculate total fuel consumed
    fuel_consumed = (total_dist / 100) * FuelConsumption_per_100km
    fuel_cost = fuel_consumed * DieselCost_per_litre

    # Calculate estimated ERP charges
    total_erp_charge = sum(float(erp['charge']) for erp in ERP_Passed)

    # Add marker layers to the map
    active_marker.add_to(Map)
    inactive_marker.add_to(Map)
    traffic_marker.add_to(Map)
    shell_marker.add_to(Map)

    # Add layer control to toggle markers
    folium.LayerControl().add_to(Map)

    map_file_path = os.path.join(app.config['MAPS_FOLDER'], 'shortest_driving_route_map.html')

    Plotted_Map.save(map_file_path)

    end_time = time.time()
    print("Execution Time: {:.2f} seconds".format(end_time - start_time))

    return render_template('viewStreetMap.html', primary_dist=route_distance,
                           total_dist=total_dist, fuel_consumed=fuel_consumed, fuel_Cost=fuel_cost,
                           erp_cost=total_erp_charge)


@app.route('/view_booking')
def view_booking():
    username = session.get('username')
    bookings = []

    csv_path = os.path.join(app.config['BOOKINGS_FOLDER'], 'guests_bookings.csv')

    # Open the CSV file and read the bookings
    with open(csv_path, mode='r') as file:
        reader = csv.DictReader(file)

        # Iterate through the rows and filter by username
        for row in reader:
            if row['Username'] == username:
                bookings.append({
                    'contact_number': row['Contact Number'],
                    'date': row['Date'],
                    'time': row['Time'],
                    'hotel': row['HotelName']
                })

    return render_template('viewBooking.html', bookings=bookings)


@app.route('/submit_route_choice', methods=['POST'])
def submit_route_choice():
    global route_type
    route_type = request.form.get('route_type')
    if route_type:
        print(f'You selected: {route_type}')
    else:
        print('Please select a route option', 'error')
    return redirect(url_for('job_assignment_map'))


if __name__ == '__main__':
    app.run(debug=True)